import openai
import config
from gtts import gTTS
import os
import random
import requests
# from pydub import AudioSegment
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from datetime import datetime

openai.api_key = config.OPENAI_API_KEY

# To generate AI response to prompt
def generate_response(prompt):

    # URL if using informed AI
    url = 'http://35.203.35.125:6000/generate_response_informed'

    # URL if using uninformed AI
    # url = "http://35.203.35.125:5000/generate_response"

    data = {"prompt":prompt}
    response = requests.post(url, json=data)

    return response.json()

# To convert text to speech
def generate_audio(text):

    tts = gTTS(text, lang='en-us')
    tts.save('audio.mp3')

    with open('audio.mp3', 'rb') as f:
        audio_content = f.read()

    os.remove('audio.mp3')

    return audio_content

# To transcribe speech to text
def transcribe(audio):
  
  audio_file = open(audio, "rb")
  transcript = openai.Audio.transcribe("whisper-1", audio_file,language = "en", response_format="json")

  return transcript.text

# To save logs of the QuickTalker
def update_log(time, latency, action, prompt, resp1, resp2):
    # Load the existing workbook or create a new one
    filename = 'logs/quicktalker.xlsx'
    workbook = openpyxl.load_workbook(filename)

    # Select the active sheet or create a new one
    sheet = workbook.active

    # Find the next available row to add the entry
    # row = sheet.max_row + 1

    # Find the next available row by searching for the first empty row below the header
    row = 2  # Start from row 2 (assuming row 1 is the header)
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    # Update the Date and Time columns with the current date and time
    now = datetime.now()
    sheet.cell(row=row, column=1, value=now.date())

    # Update the Time, Time Taken, Prompt, Response #1, Response #2 columns
    sheet.cell(row=row, column=2, value=time)
    sheet.cell(row=row, column=3, value=latency + ' s')
    sheet.cell(row=row, column=4, value=action)
    sheet.cell(row=row, column=5, value=prompt)
    sheet.cell(row=row, column=6, value=resp1)
    sheet.cell(row=row, column=7, value=resp2)

    # Save the workbook
    workbook.save(filename)

# To generate a sentence for the trainer
def generate_sentence(lineNum):

    with open("trainer_prompts.txt", "r") as file:
        lines = file.readlines()
        return lines[lineNum].strip()

# To assess the user's response to trainer
def response_score(r1, r2):

    prompt, resp   = r1[:-1].split(), r2.split()
    if r1[-1] == '.':
        prompt.append('.')
    else:
        prompt.append('?')

    wordScore = 1  # how many tries it took to get the word right
    # user gets 5 tries, anything beyond that results in a score of 0
    # each extra try costs 0.20

    scoreSum = 0  # the sum of scores for each word

    i = j = 0
    while i < len(prompt) and j < len(resp):
        if prompt[i] == resp[j]:
            i+=1
            j+=1
            scoreSum += max(wordScore, 0)
            wordScore = 1
        else:
            j+=1
            wordScore -= 0.20
    
    return round(scoreSum / len(prompt) * 100, 2)